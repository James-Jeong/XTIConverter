# -*- coding: utf-8 -*-
# ITXConverter.py
# author: jamesj

import os
import sys
from openpyxl import Workbook
import configparser

# ----------------------------------------------------------------------------------------- #
# 1) Get Parameters
iniPath = None
xlsxPath = None
sheetName = None

# 1-1) Config mode
if len(sys.argv) == 2:
	print("\n@ Config mode")
	print("- Loading config...")

	converterConfigPath = sys.argv[1]
	if not os.path.exists(converterConfigPath) or not os.path.isfile(converterConfigPath):
		print("! Fail to load the program's config file. (name=" + converterConfigPath + ")\n")
		sys.exit()

	converterConfig = configparser.ConfigParser()
	converterConfig.optionxform = lambda option: option # Preserve case for letters
	converterConfig.read(converterConfigPath, encoding="utf8")
	
	if not converterConfig.has_section("PATH"):
		print("! Fail to load the section \"PATH\". (name=" + converterConfigPath + ")\n")
		sys.exit()
	if not converterConfig.has_section("XLSX"):
		print("! Fail to load the section \"XLSX\". (name=" + converterConfigPath + ")\n")
		sys.exit()

	xlsxPath = converterConfig.get("PATH", "XLSX_PATH")
	if xlsxPath is None or len(xlsxPath) == 0:
		print("! Fail to load the option \"XLSX_PATH\" in the section \"PATH\". (name=" + converterConfigPath + ")\n")
		sys.exit()

	iniPath = converterConfig.get("PATH", "INI_PATH")
	if iniPath is None or len(iniPath) == 0:
		print("! Fail to load the option \"INI_PATH\" in the section \"PATH\". (name=" + converterConfigPath + ")\n")
		sys.exit()

	sheetName = converterConfig.get("XLSX", "SHEET_NAME")
	if sheetName is None or len(sheetName) == 0:
		print("! Fail to load the option \"SHEET_NAME\" in the section \"XLSX\". (name=" + converterConfigPath + ")\n")
		sys.exit()

# 1-2) Parameter mode
elif len(sys.argv) == 4:
	print("\n@ Parameter mode")
	print("- Loading parameters...")

	iniPath = sys.argv[1]
	if len(iniPath) == 0:
		print("! Fail to load the ini path.\n")
		sys.exit()

	xlsxPath = sys.argv[2]
	if len(xlsxPath) == 0:
		print("! Fail to load the xlsx path.\n")
		sys.exit()

	sheetName = sys.argv[3]
	if len(sheetName) == 0:
		print("! Fail to load the sheet name.\n")
		sys.exit()

else:
	print("\n! Parameter error.")
	print("argv[0]: ITXConverter.py\n")
	print("1) Config mode")
	print("argv[1]: {config path}\n")
	print("2) Parameter mode")
	print("argv[1]: {ini path}")
	print("argv[2]: {xlsx sheet name}")
	print("argv[3]: {xlsx path}\n")
	sys.exit()

print("\n@ INI Path: [ " + iniPath + " ]")
print("@ XLSX Path: [ " + xlsxPath + " ]")
print("@ Sheet Name: [ " + sheetName + " ]")

# 1-3) Check file extenstion
iniName, iniExtension = os.path.splitext(iniPath)
if iniExtension != ".ini":
	print("! INI File type is wrong. (ext=" + iniExtension + ")\n")
	sys.exit()

xlsxName, xlsxExtension = os.path.splitext(xlsxPath)
if xlsxExtension != ".xlsx":
	print("! XLSX File type is wrong. (ext=" + xlsxExtension + ")\n")
	sys.exit()

# ----------------------------------------------------------------------------------------- #
# 2) Check ini path
if not os.path.exists(iniPath) or not os.path.isfile(iniPath):
	print("! Unknown INI Path.\n")
	sys.exit()

# ----------------------------------------------------------------------------------------- #
# 3) Load ini file
print("\n- Loading ini...\n")
config = configparser.ConfigParser()
config.optionxform = lambda option: option # Preserve case for letters
config.read(iniPath, encoding="utf8")

# 3-1) Reading
sections = config.sections()
if sections is None or len(sections) == 0:
	print("! Fail. Not found any section.\n")
	sys.exit()

totalData = []
for section in sections:
	# 1] Get section
	section = str(section).strip()
	print("- [ " + section + " ]")
	options = config.options(section)
	if options is None or len(options) == 0:
		continue

	# 2] Get key & value
	data = []
	data.append(section)
	for option in options:
		keyValues = []
		key = str(option).strip()
		value = str(config.get(section, option)).strip()
		keyValues.append(key)
		keyValues.append(value)
		data.append(keyValues)
		print("	- " + key + ": " + value)

	# 3] Add section & section's data
	totalData.append(data)

# ----------------------------------------------------------------------------------------- #
# 4) Make & Write xlsx
wb = Workbook()
sheet = wb.active
sheet.title = sheetName

# 4-1) Writing
rowId = 1
for data in totalData:
	sheet.cell(row=rowId, column=1).value = data[0]
	rowId += 1
	for datum in data[1:]:
		sheet.cell(row=rowId, column=1).value = datum[0]
		sheet.cell(row=rowId, column=2).value = datum[1]
		rowId += 1

wb.save(filename=xlsxPath)

# 4-3) Check result
if os.path.exists(xlsxPath) and os.path.isfile(xlsxPath):
	print("\n@ Done.\n")
else:
	print("\n! Fail.\n")

