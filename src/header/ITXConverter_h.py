# -*- coding: utf-8 -*-
# ITXConverter_h.py
# author: jamesj

import os
import sys
from openpyxl import Workbook
import configparser

# ----------------------------------------------------------------------------------------- #
# FUNCTIONS

# @fn existProgram()
# @brief 프로그램을 종료하는 함수
def exitProgram():
	sys.exit()

# @fn checkMode(argv)
# @brief 프로그램 모드를 확인하는 함수
# 1. Config mode
# 2. Paramter mode
def checkMode(argv):
	xlsxPath = None
	sheetName = None
	iniPath = None
	argvLen = len(argv)

	# 1) Config mode
	if argvLen == 2:
		print("\n@ Config mode")
		print("- Loading config...")

		converterConfigPath = argv[1]
		if not os.path.exists(converterConfigPath) or not os.path.isfile(converterConfigPath):
			print("! Fail to load the program's config file. (name=" + converterConfigPath + ")\n")
			exitProgram()

		converterConfig = configparser.ConfigParser()
		converterConfig.optionxform = lambda option: option # Preserve case for letters
		converterConfig.read(converterConfigPath, encoding="utf8")
		
		if not converterConfig.has_section("PATH"):
			print("! Fail to load the section \"PATH\". (name=" + converterConfigPath + ")\n")
			exitProgram()
		if not converterConfig.has_section("XLSX"):
			print("! Fail to load the section \"XLSX\". (name=" + converterConfigPath + ")\n")
			exitProgram()

		xlsxPath = converterConfig.get("PATH", "XLSX_PATH")
		if xlsxPath is None or len(xlsxPath) == 0:
			print("! Fail to load the option \"XLSX_PATH\" in the section \"PATH\". (name=" + converterConfigPath + ")\n")
			exitProgram()

		iniPath = converterConfig.get("PATH", "INI_PATH")
		if iniPath is None or len(iniPath) == 0:
			print("! Fail to load the option \"INI_PATH\" in the section \"PATH\". (name=" + converterConfigPath + ")\n")
			exitProgram()

		sheetName = converterConfig.get("XLSX", "SHEET_NAME")
		if sheetName is None or len(sheetName) == 0:
			print("! Fail to load the option \"SHEET_NAME\" in the section \"XLSX\". (name=" + converterConfigPath + ")\n")
			exitProgram()

	# 2) Parameter mode
	elif argvLen == 4:
		print("\n@ Parameter mode")
		print("- Loading parameters...")

		iniPath = argv[1]
		if len(iniPath) == 0:
			print("! Fail to load the ini path.\n")
			exitProgram()

		xlsxPath = argv[2]
		if len(xlsxPath) == 0:
			print("! Fail to load the xlsx path.\n")
			exitProgram()

		sheetName = argv[3]
		if len(sheetName) == 0:
			print("! Fail to load the sheet name.\n")
			exitProgram()

	else:
		print("\n! Parameter error.")
		print("argv[0]: ITXConverter.py\n")
		print("1) Config mode")
		print("argv[1]: {config path}\n")
		print("2) Parameter mode")
		print("argv[1]: {ini path}")
		print("argv[2]: {xlsx sheet name}")
		print("argv[3]: {xlsx path}\n")
		exitProgram()

	return xlsxPath, sheetName, iniPath

# @fn checkXlsx(xlsxPath)
# @brief 지정한 XLSX 파일의 확장자를 검사하는 함수
def checkXlsx(xlsxPath):
	xlsxName, xlsxExtension = os.path.splitext(xlsxPath)
	if xlsxExtension != ".xlsx":
	        print("! XLSX File type is wrong. (ext=" + xlsxExtension + ")\n")
        	exitProgram()

# @fn checkIni(iniPath)
# @brief 지정한 INI 경로 존재 여부와 파일의 확장자를 검사하는 함수
def checkIni(iniPath):
	if not os.path.exists(iniPath) or not os.path.isfile(iniPath):
		print("! Unknown INI Path.\n")
		exitProgram()

	iniName, iniExtension = os.path.splitext(iniPath)
	if iniExtension != ".ini":
	        print("! INI File type is wrong. (ext=" + iniExtension + ")\n")
	        exitProgram()

# @fn loadIni(iniPath)
# @brief 지정한 INI 파일 로드하여 저장된 값들을 반환하는 함수
def loadIni(iniPath):
	print("\n- Loading ini...\n")
	config = configparser.ConfigParser()
	config.optionxform = lambda option: option # Preserve case for letters
	config.read(iniPath, encoding="utf8")

	sections = config.sections()
	if sections is None or len(sections) == 0:
		print("! Fail. Not found any section.\n")
		exitProgram()

	totalData = []
	for section in sections:
		# 1] Get section
		section = str(section).strip()
		print("- [ " + section + " ]")
		options = config.options(section)
		if options is None or len(options) == 0:
			continue

		# 2] Get key & value
		data = []
		data.append(section)
		for option in options:
			keyValues = []
			key = str(option).strip()
			value = str(config.get(section, option)).strip()
			keyValues.append(key)
			keyValues.append(value)
			data.append(keyValues)
			print("	- " + key + ": " + value)

		# 3] Add section & section's data
		totalData.append(data)

	return totalData

# @fn writeXlsx(xlsxPath, sheetName, totalData)
# @brief 지정한 XLSX 경로와 Sheet 이름, 저장된 값들을 통해 새로운 XLSX 파일을 생성하는 함수
def writeXlsx(xlsxPath, sheetName, totalData):
	wb = Workbook()
	sheet = wb.active
	sheet.title = sheetName
	
	rowId = 1
	for data in totalData:
		sheet.cell(row=rowId, column=1).value = data[0]
		rowId += 1
		for datum in data[1:]:
			sheet.cell(row=rowId, column=1).value = datum[0]
			sheet.cell(row=rowId, column=2).value = datum[1]
			rowId += 1

	wb.save(filename=xlsxPath)
	
	if os.path.exists(xlsxPath) and os.path.isfile(xlsxPath):
		print("\n@ Done.\n")
	else:
		print("\n! Fail.\n")


